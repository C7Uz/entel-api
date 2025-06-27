from typing import Optional
from django.contrib import admin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.utils.translation import gettext_lazy as _
from unfold.admin import ModelAdmin
from django import forms
from django.contrib.auth.forms import UserCreationForm, UserChangeForm


# Import / Export
from import_export.admin import ImportExportModelAdmin
from import_export.resources import ModelResource
from unfold.contrib.import_export.forms import ExportForm, ImportForm
from import_export import fields

# Models
from .models import User, Inscriptions



from django.http import HttpResponse
from openpyxl import Workbook # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
from datetime import datetime

# Solo agregar este mixin antes de tu clase InscriptionsAdmin existente

# Solo agregar este mixin antes de tu clase InscriptionsAdmin existente
class ExcelExportMixin:
    def export_to_excel_custom(self, request, queryset):
        """Exportación personalizada a Excel con estilos"""
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Inscripciones"
        
        # Configurar encabezados
        headers = [
            'ID', 'Evento', 'Nombre Completo', 'Teléfono', 'Email',
            'Trabajo', 'Tipo', 'Empresa', 'RUC', 'Publicidad', 'Fecha Inscripción'
        ]
        ws.append(headers)
        
        # Estilo para encabezados
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Aplicar estilo a encabezados
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = ws[f'{col_letter}1']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Agregar datos
        for inscription in queryset:
            data_row = [
                inscription.id,
                inscription.event.title if inscription.event else '',
                inscription.fullname,
                inscription.cellphone,
                inscription.email,
                inscription.job,
                "Empresa" if inscription.flag_business else "Particular",
                inscription.company or '',
                inscription.ruc or '',
                "Sí" if inscription.publicidad else "No",
                inscription.created.strftime('%d/%m/%Y %H:%M') if inscription.created else ''
            ]
            ws.append(data_row)
        
        # Ajustar ancho de columnas
        column_widths = [8, 25, 20, 15, 25, 20, 12, 20, 15, 12, 18]
        for col_num, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = width
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"inscripciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar y retornar
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        response.write(excel_file.getvalue())
        
        return response
    
    export_to_excel_custom.short_description = "Exportar seleccionados a Excel (Personalizado)"



# Resource
class InscriptionResource(ModelResource):
    event = fields.Field(attribute='event__title', column_name="Evento")
    
    class Meta:
        model = Inscriptions
        fields = ['event', 'fullname', 'cellphone', 'job', 'flag_business', 'company', 'ruc', 'email','publicidad', 'created']

 

class CustomUserCreationForm(UserCreationForm):
    class Meta:
        model = User
        fields = ('email', 'first_name', 'last_name')


class CustomUserChangeForm(forms.ModelForm):
    password = forms.CharField(
        label=_("Nueva contraseña"),
        widget=forms.PasswordInput,
        required=False,
        help_text=_("Dejar en blanco para no cambiar la contraseña.")
    )

    class Meta:
        model = User
        fields = ('email', 'first_name', 'last_name', 'password', 'is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')

    def save(self, commit=True):
        user = super().save(commit=False)
        password = self.cleaned_data.get('password')
        if password:
            user.set_password(password)  # Aquí se encripta
        if commit:
            user.save()
            self.save_m2m()
        return user


@admin.register(User)
class UserAdmin(BaseUserAdmin):
    form = CustomUserChangeForm
    add_form = CustomUserCreationForm

    list_display = ('id', 'first_name', 'last_name', 'email')
    ordering = ['first_name']
    list_per_page = 50
    filter_horizontal = ('groups', 'user_permissions')

    fieldsets = (
        (_('Información Personal'), {'fields': ('first_name', 'last_name', 'email', 'password')}),
        (_('Permisos'), {
            'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions'),
        }),
    )
    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('email', 'first_name', 'last_name', 'password1', 'password2'),
        }),
    )

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(Inscriptions)
class InscriptionsAdmin(ModelAdmin, ImportExportModelAdmin, ExcelExportMixin):
    list_display = ('id', 'event', 'fullname','publicidad', 'created')
    list_filter = ('event__title', 'created', 'flag_business', 'response_infobip')
    search_fields = ('event__title', 'fullname', 'email', 'job', 'ruc')
    list_per_page = 50
    import_form_class = ImportForm
    export_form_class = ExportForm
    resource_class = InscriptionResource
    actions = ['export_to_excel_custom']
    def has_import_permission(self, request):
        return False
    
    def has_add_permission(self, request):
        return False
    
    def has_change_permission(self, request, obj=None):
        return False
    
    def has_delete_permission(self, request, obj=None):
        return True
    
